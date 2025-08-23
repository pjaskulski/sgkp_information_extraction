""" uzupełnienie powiatów ujednoliconych na podstawie nazwy powiatu z OCR i dostarczonej tabeli powiatów """
import os
import json
import time
from pathlib import Path
import openpyxl


VOLUMES = ['01', '02', '03', '04', '05', '06', '07', '08',
           '09', '10', '11', '12', '13', '14', '15', '16']

# -------------- Główny program ------------------------------------------------
if __name__ == '__main__':

     # pomiar czasu wykonania
    start_time = time.time()

    powiaty_path = Path('..') / 'dictionary' / 'powiaty_ujednolicone.xlsx'

    # przetwarzanie arkusza
    wb = openpyxl.load_workbook(powiaty_path)
    ws = wb.active

    # kolumny w źródłowym pierwszym arkuszu
    col_names = {}
    nr_col = 0
    for column in ws.iter_cols(1, ws.max_column):
        col_names[column[0].value] = nr_col
        nr_col += 1

    max_row = ws.max_row

    powiaty = {}

    for row in ws.iter_rows(min_row=2, max_row=max_row):
        numer = row[0].value
        nazwa_json = str(row[col_names["nazwa_json"]].value).strip()
        nazwa_ujednolicona = str(row[col_names["nazwa_ujednolicona"]].value).strip()
        powiaty[nazwa_json] = nazwa_ujednolicona

    # przetwarzanie całej kolekcji tomów
    for VOLUME in VOLUMES:
        input_path = Path('..') / 'SGKP' / 'JSON' / 'merged_dane_podstawowe' / f'sgkp_{VOLUME}_dane_podstawowe.json'
        output_path = Path('..') / 'SGKP' / 'JSON' / 'merged_dane_podstawowe' / f'sgkp_{VOLUME}_powiaty.json'

        if os.path.exists(output_path):
            os.remove(output_path)

        with open(input_path, "r", encoding='utf-8') as f:
            json_data = json.load(f)

            for item_result in json_data:
                rodzaj = item_result["rodzaj"]

                if rodzaj == "zbiorcze":
                    elementy = item_result["elementy"]
                    for element in elementy:
                        powiat = element.get("powiat_ocr", None)
                        if powiat and powiat in powiaty:
                            element["powiat_ujednolicony"] = powiaty[powiat]
                else:
                    powiat = item_result.get("powiat_ocr", None)
                    if powiat and powiat in powiaty:
                        item_result["powiat_ujednolicony"] = powiaty[powiat]

        # zapis
        with open(output_path, 'w', encoding='utf-8') as f_out:
            json.dump(json_data, f_out, indent=4, ensure_ascii=False)

     # czas wykonania programu
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f'Czas wykonania programu: {time.strftime("%H:%M:%S", time.gmtime(elapsed_time))} s.')
